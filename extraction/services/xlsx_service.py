from __future__ import annotations
from templates.models import Template, TemplateField
from extraction.models import ExtractionBatch
from extraction.services.base import BaseInvoiceExtractionService, ExtractionProcessingError, _stringify_cell
import openpyxl


class InvoiceXlsxExtractionService(BaseInvoiceExtractionService):
    file_format = ExtractionBatch.FileFormat.XLSX

    def __init__(self, template: Template, supplier_catalog=None):
        if template.document_type != Template.DocumentType.XLSX:
            raise ExtractionProcessingError("El template seleccionado no es de tipo XLSX.")
        super().__init__(template, supplier_catalog)

    def _load_template_fields(self):
        return (self.template.fields.select_related("layout_field")
                .filter(extraction_type=TemplateField.ExtractionType.HEADER_NAME))

    def _load_headers(self, workbook) -> dict:
        sheet = workbook.worksheets[0]
        headers: dict[str, list[int]] = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value is None:
                continue
            headers.setdefault(str(cell.value), []).append(col_idx)
        return headers

    def _iter_source_units(self, uploaded_file):
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = workbook.worksheets[0]
        headers = self._load_headers(workbook)
        for row_idx in range(2, sheet.max_row + 1):
            row_values, all_blank = {}, True
            for tf in self.template_fields:
                occurrences = headers.get(tf.source_field.strip(), [])
                occurrence_idx = (tf.header_occurrence or 1) - 1
                col_idx = occurrences[occurrence_idx] if 0 <= occurrence_idx < len(occurrences) else None
                value = sheet.cell(row=row_idx, column=col_idx).value if col_idx is not None else None
                if value is not None:
                    all_blank = False
                row_values[tf.id] = _stringify_cell(value)
            if not all_blank:
                yield (row_idx, row_values)