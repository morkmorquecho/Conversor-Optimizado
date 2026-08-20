"""
Servicio de extracción de facturas XLSX.
Flujo:
1. Se recibe un archivo Excel, un Template XLSX y opcionalmente un SupplierCatalog.
2. Se detectan las filas de datos y, por cada fila, se crea un ExtractionJob dentro de un ExtractionBatch.
3. Se extraen los campos registrados en TemplateField por nombre de encabezado (soportando columnas duplicadas mediante header_occurrence).
4. Se aplican las reglas de normalización configuradas para cada TemplateField.
5. Se guardan los valores extraídos como ExtractionResult.
6. Si existe un SupplierCatalog:
   - Se obtiene el valor pivote desde el resultado extraído del Excel.
   - Se busca la fila correspondiente dentro de SupplierCatalogRow.
   - El pivot_field_name del SupplierCatalog únicamente define qué campo del catálogo se utiliza como referencia.
   - No es necesario que el campo pivote esté registrado como SupplierCatalogColumn.
7. Una vez encontrada la fila del catálogo, se extraen únicamente las columnas configuradas mediante SupplierCatalogColumnLayoutField.
8. Se resuelven los campos de sistema.
9. Se genera un Excel de salida con todas las columnas del Layout, una fila por ExtractionJob.
"""
from __future__ import annotations
import re
from datetime import datetime
from io import BytesIO
from typing import Optional
import openpyxl
from django.apps import apps
from django.utils import timezone
from catalogs.models import SupplierCatalog, SupplierCatalogColumnLayoutField, SupplierCatalogPivotMapping
from extraction.models import ExtractionBatch, ExtractionError as ExtractionErrorModel, ExtractionJob, ExtractionResult
from layouts.models import LayoutField, NormalizationRule
from templates.models import Template, TemplateField
from layouts.system_fields import SYSTEM_FIELD_REGISTRY
class ExtractionProcessingError(Exception):
    """Error de configuración o de datos que impide continuar el proceso."""

def _stringify_cell(value) -> str:
    """Convierte un valor de una celda Excel a string."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()

def apply_normalization_rule(rule: NormalizationRule, value: str) -> str:
    """Aplica una sola regla de normalización sobre un valor."""
    if value in (None, ""):
        return value
    config = rule.config or {}
    if rule.rule_type == NormalizationRule.RuleType.TRIM:
        return value.strip()
    if rule.rule_type == NormalizationRule.RuleType.UPPERCASE:
        return value.strip().upper()
    if rule.rule_type == NormalizationRule.RuleType.REGEX_REPLACE:
        pattern = config.get("pattern", "")
        replacement = config.get("replacement", "")
        if not pattern:
            return value
        return re.sub(pattern, replacement, value)
    if rule.rule_type == NormalizationRule.RuleType.DATE_FORMAT:
        input_format = config.get("input_format")
        output_format = config.get("output_format")
        if not input_format or not output_format:
            return value
        try:
            parsed = datetime.strptime(value.strip(), input_format)
        except ValueError:
            return value
        return parsed.strftime(output_format)
    if rule.rule_type == NormalizationRule.RuleType.VALUE_MAP:
        return _apply_value_map(config, value)
    return value

def _apply_value_map(config: dict, value: str) -> str:
    """
    VALUE_MAP extendido.
    config = {
        "map": {"DLS": "USA"},
        "case_insensitive": true,
        "default": null,
        "lookup": {
            "app_label": "catalogs",
            "model": "Currency",
            "match_field": "country",
            "result_field": "code",
        }
    }
    """
    mapping = config.get("map", {}) or {}
    case_insensitive = config.get("case_insensitive", True)
    key = value.strip()
    if case_insensitive:
        mapping = {k.upper(): v for k, v in mapping.items()}
        key = key.upper()
    if key in mapping:
        mapped_value = mapping[key]
    else:
        mapped_value = config.get("default", value)
    lookup_cfg = config.get("lookup")
    if not lookup_cfg:
        return mapped_value
    try:
        Model = apps.get_model(lookup_cfg["app_label"], lookup_cfg["model"])
        obj = Model.objects.get(**{lookup_cfg["match_field"]: mapped_value})
        return str(getattr(obj, lookup_cfg.get("result_field", "code")))
    except Exception:
        return mapped_value

def apply_normalization_chain(template_field: TemplateField, raw_value: str) -> str:
    """Aplica todas las reglas de normalización de un TemplateField respetando su orden."""
    value = raw_value
    rules = template_field.rules.select_related("normalization_rule").order_by("sort_order")
    for template_field_rule in rules:
        value = apply_normalization_rule(template_field_rule.normalization_rule, value)
    return value


class InvoiceXlsxExtractionService:
    def __init__(self, template: Template, supplier_catalog: Optional[SupplierCatalog] = None):
        if template.document_type != Template.DocumentType.XLSX:
            raise ExtractionProcessingError("El template seleccionado no es de tipo XLSX.")
        self.template = template
        self.supplier = template.supplier
        self.layout = template.layout
        self.supplier_catalog = supplier_catalog
        self.template_fields = list(
            template.fields
            .select_related("layout_field")
            .filter(extraction_type=TemplateField.ExtractionType.HEADER_NAME)
        )
        if not self.template_fields:
            raise ExtractionProcessingError("El template no tiene campos configurados para extracción por encabezado.")
        self.layout_fields = list(self.layout.fields.order_by("sort_order"))
        self._catalog_mappings = []
        self._pivot_mapping = None
        if self.supplier_catalog is not None:
            self._catalog_mappings = self._resolve_catalog_mappings()
            self._pivot_mapping = (
                SupplierCatalogPivotMapping.objects
                .filter(template=self.template, supplier_catalog=self.supplier_catalog)
                .select_related("pivot_template_field__layout_field")
                .first()
            )

    def _resolve_catalog_mappings(self):
        """
        Obtiene únicamente las columnas del catálogo configuradas para
        extraerse hacia el Layout. El pivote se resuelve por separado,
        vía SupplierCatalogPivotMapping (ver __init__).
        """
        return list(
            SupplierCatalogColumnLayoutField.objects.filter(
                column__supplier_catalog=self.supplier_catalog,
                layout_field__layout=self.layout,
            ).select_related("column", "layout_field")
        )

    def _load_headers(self, workbook) -> dict:
        """
        Encabezados de la primera fila de la PRIMERA hoja del workbook,
        sin importar cómo se llame la hoja.

        CAMBIO: Ahora retorna un diccionario donde la clave es el nombre del
        encabezado y el valor es una LISTA de los índices de columna donde aparece.
        Esto permite manejar encabezados duplicados mediante header_occurrence.
        """
        sheet = workbook.worksheets[0]
        headers = {}

        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value is None:
                continue
            header_name = str(cell.value)
            if header_name not in headers:
                headers[header_name] = []
            
            headers[header_name].append(col_idx)

        return headers

    def _iter_data_rows(self, workbook):
        """
        Itera las filas de datos del Excel desde la PRIMERA hoja.
        Cada fila genera: (row_number, raw_values_by_tf_id)
        """
        sheet = workbook.worksheets[0]
        headers = self._load_headers(workbook)
        
        for row_idx in range(2, sheet.max_row + 1):
            row_values = {}
            all_blank = True
            
            for tf in self.template_fields:
                header_name = tf.source_field.strip()
                occurrences = headers.get(header_name, [])
                
                # CAMBIO: Usamos header_occurrence para seleccionar el índice correcto.
                # Restamos 1 porque el campo en BD es 1-indexed y las listas en Python son 0-indexed.
                occurrence_idx = (tf.header_occurrence or 1) - 1 
                
                col_idx = None
                if 0 <= occurrence_idx < len(occurrences):
                    col_idx = occurrences[occurrence_idx]
                
                value = None
                if col_idx is not None:
                    value = sheet.cell(row=row_idx, column=col_idx).value
                
                if value is not None:
                    all_blank = False
                row_values[tf.id] = _stringify_cell(value)
            
            if not all_blank:
                yield (row_idx, row_values)

    def process(self, uploaded_file, source_file_name: str):
        """
        Procesa un archivo XLSX completo.
        Crea: 1 ExtractionBatch, N ExtractionJobs
        Un ExtractionJob representa una fila del Excel.
        Retorna: (output_xlsx_bytes, extraction_batch)
        """
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        rows = list(self._iter_data_rows(workbook))
        if not rows:
            raise ExtractionProcessingError("No se encontraron filas de datos en el excel para el template seleccionado.")
        batch = ExtractionBatch.objects.create(
            supplier=self.supplier,
            source_file=source_file_name,
            file_format=ExtractionBatch.FileFormat.XLSX,
            status=ExtractionBatch.Status.PENDING,
            template=self.template,
            supplier_catalog=self.supplier_catalog,
            total_records=len(rows),
        )
        jobs = []
        for row_idx, raw_values_by_tf_id in rows:
            job = self._process_row(batch=batch, row_number=row_idx, raw_values_by_tf_id=raw_values_by_tf_id)
            jobs.append(job)
        successful_records = sum(1 for job in jobs if job.status == ExtractionJob.Status.PROCESSED)
        failed_records = sum(1 for job in jobs if job.status == ExtractionJob.Status.REVIEW)
        if failed_records > 0:
            batch.status = ExtractionBatch.Status.REVIEW
        else:
            batch.status = ExtractionBatch.Status.PROCESSED
        batch.successful_records = successful_records
        batch.failed_records = failed_records
        batch.processed_at = timezone.now()
        batch.save(update_fields=["status", "successful_records", "failed_records", "processed_at"])
        output_bytes = self._build_output_workbook(jobs)
        return (output_bytes, batch)

    def _process_row(self, batch: ExtractionBatch, row_number: int, raw_values_by_tf_id: dict) -> ExtractionJob:
        job = ExtractionJob.objects.create(
            extraction_batch=batch,
            row_number=row_number,
            status=ExtractionJob.Status.PENDING,
        )
        had_errors = False
        # 1. Extraer campos directamente desde el Excel
        for tf in self.template_fields:
            raw_value = raw_values_by_tf_id.get(tf.id, "")
            normalized_value = apply_normalization_chain(tf, raw_value)
            ExtractionResult.objects.update_or_create(
                extraction_job=job,
                layout_field=tf.layout_field,
                defaults={"raw_value": raw_value, "normalized_value": normalized_value},
            )
        # 2. Buscar datos complementarios en el catálogo
        if self.supplier_catalog is not None:
            catalog_success = self._fill_from_catalog(job)
            if not catalog_success:
                had_errors = True
        # 3. Campos de sistema
        self._fill_system_fields(job, raw_values_by_tf_id)
        # 4. Estado final del Job
        job.status = ExtractionJob.Status.REVIEW if had_errors else ExtractionJob.Status.PROCESSED
        job.processed_at = timezone.now()
        job.save(update_fields=["status", "processed_at"])
        return job

    def _fill_from_catalog(self, job: ExtractionJob) -> bool:
        """
        Busca una fila del catálogo utilizando el valor pivote que ya fue extraído desde el Excel.
        El pivot_field_name del SupplierCatalog indica la referencia utilizada para identificar la fila correspondiente.
        Una vez encontrada la fila:
            - NO se extrae el campo pivote.
            - Solo se extraen las columnas configuradas mediante SupplierCatalogColumnLayoutField.
        """
        if self._pivot_mapping is None:
            ExtractionErrorModel.objects.create(
                extraction_job=job,
                message=(
                    f"No hay configuración de pivote para el template "
                    f"'{self.template.name}' y catálogo '{self.supplier_catalog.name}'."
                ),
            )
            return False
        pivot_layout_field = self._pivot_mapping.pivot_template_field.layout_field
        pivot_result = ExtractionResult.objects.filter(
            extraction_job=job,
            layout_field=pivot_layout_field,
        ).first()
        pivot_value = ""
        if pivot_result:
            pivot_value = pivot_result.normalized_value or pivot_result.raw_value
        if not pivot_value:
            ExtractionErrorModel.objects.create(
                extraction_job=job,
                layout_field=pivot_layout_field.layout_field,
                message="No se pudo obtener el valor pivote desde el Excel para consultar el catálogo.",
            )
            return False
        catalog_row = self.supplier_catalog.rows.filter(pivot_value=pivot_value).first()
        if catalog_row is None:
            ExtractionErrorModel.objects.create(
                extraction_job=job,
                layout_field=pivot_layout_field.layout_field,
                message=f"No se encontró una fila en el catálogo '{self.supplier_catalog.name}' para el valor pivote '{pivot_value}'.",
            )
            return False
        for mapping in self._catalog_mappings:
            value = _stringify_cell(catalog_row.data.get(mapping.column.source_name))
            ExtractionResult.objects.update_or_create(
                extraction_job=job,
                layout_field=mapping.layout_field,
                defaults={"raw_value": value, "normalized_value": value},
            )
        return True

    def _fill_system_fields(self, job: ExtractionJob, raw_values_by_tf_id: dict):
        """Resuelve campos calculados por el sistema."""
        for layout_field in self.layout_fields:
            if not layout_field.system_field_key:
                continue
            handler = SYSTEM_FIELD_REGISTRY.get(layout_field.system_field_key)
            if handler is None:
                continue
            value = _stringify_cell(handler(job, raw_values_by_tf_id))
            ExtractionResult.objects.update_or_create(
                extraction_job=job,
                layout_field=layout_field,
                defaults={"raw_value": value, "normalized_value": value},
            )

    def _build_output_workbook(self, jobs) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.layout.code[:31] or "output"
        # Headers
        ws.append([field.name for field in self.layout_fields])
        # Una fila por ExtractionJob
        for job in jobs:
            results_by_field_id = {
                result.layout_field_id: (result.normalized_value or result.raw_value)
                for result in job.results.all()
            }
            ws.append([results_by_field_id.get(field.id, "") for field in self.layout_fields])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()